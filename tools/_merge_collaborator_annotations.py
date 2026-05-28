"""
Merge team annotations from the filled xlsx into canonical JSON under a new
top-level `collaborator_annotations` section. Also folds LLM sheet decisions
into a parallel `llm_collaborator_input` section and updates one entry in
`internal_decisions_needed` (D61 → 문서 버전 전략).

Run with --dry-run to print the proposed canonical JSON diff without writing.
Run with --apply to actually write canonical JSON.
"""
import argparse
import io
import json
import sys
from copy import deepcopy
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
CANONICAL = ROOT / "spec_canonical" / "backend_spec_v3.canonical.json"
NEW_XLSX = ROOT / "outputs" / "snail_백엔드_협업명세서_v3_20260527_1825.xlsx"
DIFF = ROOT / "outputs" / "_sheet_diff.json"

SHEET_ENTITIES = {
    "3.유저(앱)_회원": ["User"],
    "5.사장님(웹)_샵관리": ["Owner", "Shop", "Designer"],
    "6.사장님(웹)_디자인": ["Design", "DesignImage"],
    "8.커뮤니티_스네일": ["Snap"],
    "9.커뮤니티_댓글": ["Comment"],
    "10.커뮤니티_리뷰": ["Review"],
    "11.커뮤니티_신고": ["Report"],
}

def detect_entity_for_row(ws, row, candidate_entities):
    """Walk upward from `row` to find the closest section header like
    '<EntityName> 필드 정의' and return that entity if it is in candidates."""
    for r in range(row, 0, -1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        s = str(v).strip()
        for ent in candidate_entities:
            if s == f"{ent} 필드 정의" or s.startswith(f"{ent} "):
                return ent
    return candidate_entities[0]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="write canonical JSON")
    args = ap.parse_args()

    with CANONICAL.open(encoding="utf-8") as f:
        canonical = json.load(f)
    original = deepcopy(canonical)

    wb = load_workbook(NEW_XLSX, data_only=True)
    diff = json.loads(DIFF.read_text(encoding="utf-8"))

    def is_in_diff(sheet_name, coord):
        return coord in diff["sheets"].get(sheet_name, {}).get("added", {})

    # ---- 1. Build collaborator_annotations.entities ----
    entity_annotations = {}
    freeform = {}
    summary_entries = []

    for sheet_name, candidates in SHEET_ENTITIES.items():
        ws = wb[sheet_name]
        sheet_added = diff["sheets"].get(sheet_name, {}).get("added", {})
        for row in range(4, ws.max_row + 1):
            field_cell = ws.cell(row, 1).value
            note_cell = ws.cell(row, 8).value  # column H
            extra_cell = ws.cell(row, 9).value  # column I (rare, e.g., 3.H12 has I12)
            # If column A is newly added in diff and there's no field-row note,
            # treat it as a freeform note (only A-column additions count)
            if is_in_diff(sheet_name, f"A{row}") and field_cell:
                freeform.setdefault(sheet_name, []).append({
                    "cell": f"A{row}", "text": str(field_cell).strip()
                })
                continue
            if not field_cell:
                continue
            field_name = str(field_cell).strip()
            # Only process this row if column H or I was newly added by the team
            h_added = is_in_diff(sheet_name, f"H{row}")
            i_added = is_in_diff(sheet_name, f"I{row}")
            if not (h_added or i_added):
                continue
            if note_cell is None and extra_cell is None:
                continue
            entity = detect_entity_for_row(ws, row, candidates)
            slot = entity_annotations.setdefault(entity, {})
            note_text = str(note_cell).strip() if note_cell else ""
            if extra_cell:
                if note_text:
                    note_text = note_text + "\n" + str(extra_cell).strip()
                else:
                    note_text = str(extra_cell).strip()
            if note_text:
                # If same field appears twice, append
                if field_name in slot and slot[field_name] != note_text:
                    slot[field_name] = slot[field_name] + " | " + note_text
                else:
                    slot[field_name] = note_text
                summary_entries.append((sheet_name, entity, field_name, note_text))

    # ---- 2. LLM sheet (12) ----
    ws = wb["12.LLM명세"]
    llm_input = {"transform": {}, "classify": {}, "standard_tags": {}, "error_codes": {}}
    # Transform: rows 10-14, col C; key derived from col A
    transform_rows = {10: "purpose", 11: "suggested_endpoint", 12: "request_fields",
                       13: "response_fields", 14: "recommendation"}
    for r, key in transform_rows.items():
        v = ws.cell(r, 3).value
        if v:
            llm_input["transform"][key] = str(v).strip()
    classify_rows = {19: "purpose", 20: "suggested_endpoint", 21: "request_fields",
                      22: "response_fields"}
    for r, key in classify_rows.items():
        v = ws.cell(r, 3).value
        if v:
            llm_input["classify"][key] = str(v).strip()
    # Standard tags: rows 27-33, key in col A, value in col C
    for r in range(27, 34):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 3).value
        if k and v:
            llm_input["standard_tags"][str(k).strip()] = str(v).strip()
    # Error codes: rows 38-43, code in col A, modification in col D
    for r in range(38, 44):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 4).value
        if k and v:
            llm_input["error_codes"][str(k).strip()] = str(v).strip()

    # ---- 3. Decisions sheet (14) D61 → internal_decisions_needed[0] ----
    ws = wb["14.의사결정기록"]
    # Map by topic name to be robust
    decisions_updated = []
    for r in range(60, 80):
        topic = ws.cell(r, 1).value
        answer = ws.cell(r, 4).value
        if topic and answer:
            topic_s = str(topic).strip()
            ans_s = str(answer).strip()
            if topic_s in ("주제",):
                continue
            for item in canonical.get("internal_decisions_needed", []):
                if item.get("topic") == topic_s:
                    item["answer"] = ans_s
                    decisions_updated.append((topic_s, ans_s))
                    break

    # ---- 4. Inject into canonical ----
    canonical["collaborator_annotations"] = {
        "_meta": {
            "source_file": NEW_XLSX.name,
            "imported_at": "2026-05-27",
            "cell_count": len(summary_entries) + sum(len(v) for v in freeform.values()),
        },
        "entities": entity_annotations,
        "freeform_notes": freeform,
    }
    canonical["llm_collaborator_input"] = llm_input

    # ---- 5. Print summary ----
    print("=" * 80)
    print("MERGE PLAN")
    print("=" * 80)
    print(f"\n[entities]  {sum(len(v) for v in entity_annotations.values())} field annotations across {len(entity_annotations)} entities")
    for ent, fields in entity_annotations.items():
        print(f"  • {ent}: {len(fields)} fields")
        for fname, txt in fields.items():
            disp = txt.replace("\n", " ⏎ ")[:90]
            print(f"      - {fname:25} → {disp}")
    if freeform:
        print(f"\n[freeform_notes]")
        for sh, notes in freeform.items():
            for n in notes:
                print(f"  • {sh} {n['cell']}: {n['text']}")
    print(f"\n[llm_collaborator_input]")
    for section, fields in llm_input.items():
        if not fields: continue
        print(f"  • {section}: {len(fields)} items")
        for k, v in fields.items():
            print(f"      - {k:25} → {v[:80].replace(chr(10),' ⏎ ')}")
    print(f"\n[internal_decisions_needed]  {len(decisions_updated)} item(s) gained `answer` field")
    for topic, ans in decisions_updated:
        print(f"  • {topic} → answer: {ans}")

    if args.apply:
        with CANONICAL.open("w", encoding="utf-8") as f:
            json.dump(canonical, f, ensure_ascii=False, indent=2)
            f.write("\n")
        print(f"\n✅ Wrote canonical JSON: {CANONICAL}")
    else:
        print(f"\n(dry-run — pass --apply to write canonical JSON)")

if __name__ == "__main__":
    main()
