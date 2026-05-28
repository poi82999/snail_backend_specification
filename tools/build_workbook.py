import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
CANONICAL_PATH = ROOT / "spec_canonical" / "backend_spec_v3.canonical.json"
SPEC_TEXT_DIR = ROOT / "spec_text"
OUTPUT_PATH = ROOT / "outputs" / "네일예약_백엔드_협업명세서_v3.xlsx"

FN = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FN, bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", start_color="D9E1F2")
SECTION_FONT = Font(name=FN, bold=True, color="1F4E78", size=11)
REQUIRED_FILL = PatternFill("solid", start_color="FFF2CC")
NOTE_FILL = PatternFill("solid", start_color="E2EFDA")
WHO_FILL = PatternFill("solid", start_color="FCE4D6")
WARN_FILL = PatternFill("solid", start_color="F8CBAD")

TITLE_FONT = Font(name=FN, bold=True, size=16, color="1F4E78")
SUB_FONT = Font(name=FN, bold=True, size=12, color="305496")
BODY = Font(name=FN, size=10)
BODY_B = Font(name=FN, size=10, bold=True)

thin = Side(border_style="thin", color="B4C7E7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TL = Alignment(horizontal="left", vertical="top", wrap_text=True)

DEFAULT_COL_WIDTH = 12
MAX_COL_WIDTH = 72
MIN_ROW_HEIGHT = 28
MAX_ROW_HEIGHT = 260
LINE_HEIGHT = 15
ROW_PADDING = 10

COMMON_FIELD_MEANINGS = {
    "user_id": "앱 유저 한 명을 서버가 구분하기 위한 내부 ID입니다. 화면에 직접 보여주지 않습니다.",
    "owner_id": "사장님 계정 한 명을 서버가 구분하기 위한 내부 ID입니다. 사장님 웹 권한 확인에 사용합니다.",
    "shop_id": "네일샵 하나를 서버가 구분하기 위한 내부 ID입니다. 샵 상세, 예약, 디자인, 리뷰를 연결합니다.",
    "designer_id": "디자이너 한 명을 서버가 구분하기 위한 내부 ID입니다. 예약 배정과 스케줄 계산에 사용합니다.",
    "design_id": "예약 가능한 네일 디자인 상품 하나를 구분하는 내부 ID입니다.",
    "reservation_id": "예약 한 건을 구분하는 내부 ID입니다. 리뷰, 알림, 취소/노쇼 이력과 연결됩니다.",
    "payment_id": "향후 결제 기능에서 사용할 결제 ID입니다. MVP 예약 명세에는 기본 노출하지 않습니다.",
    "refund_id": "향후 환불 기능에서 사용할 환불 ID입니다. MVP 예약 명세에는 기본 노출하지 않습니다.",
    "image_id": "업로드된 이미지 한 장을 구분하는 내부 ID입니다.",
    "snap_id": "커뮤니티 스네일(Snap) 게시글 한 건을 구분하는 내부 ID입니다.",
    "comment_id": "댓글 한 건을 구분하는 내부 ID입니다.",
    "review_id": "리뷰 한 건을 구분하는 내부 ID입니다.",
    "report_id": "신고 한 건을 구분하는 내부 ID입니다.",
    "created_at": "데이터가 처음 생성된 시각입니다. 화면에서는 작성일/등록일/가입일로 표시될 수 있습니다.",
    "status": "현재 처리 상태입니다. 상태에 따라 화면 문구, 버튼 노출, 수정 가능 여부가 달라집니다.",
    "name": "사람, 샵, 디자이너 등 화면에 표시되는 이름입니다.",
    "title": "디자인이나 콘텐츠를 유저에게 보여줄 때 쓰는 제목입니다.",
    "description": "샵 또는 디자인을 설명하는 긴 문구입니다.",
    "phone": "연락처입니다. 알림톡 수신 또는 샵 문의에 사용될 수 있습니다.",
    "email": "로그인과 계정 식별에 사용하는 이메일입니다.",
    "password": "사장님 웹 로그인에 사용하는 비밀번호입니다. 서버에는 원문이 아니라 해시로 저장합니다.",
    "nickname": "앱 커뮤니티와 프로필에 표시되는 유저 이름입니다.",
    "profile_image_url": "프로필 사진의 저장 URL입니다. 실제 파일 업로드 후 서버가 URL을 저장합니다.",
    "thumbnail_url": "목록 카드에서 대표로 보여줄 이미지 URL입니다.",
    "image_urls": "여러 장의 이미지 URL 목록입니다. 업로드 순서와 최대 장수를 화면에서 제한해야 합니다.",
    "tags": "검색, 분류, 필터에 사용하는 태그 목록입니다. 자유 입력인지 표준 태그 선택인지 확인이 필요합니다.",
    "like_count": "좋아요 수입니다. 서버가 자동 계산하며 화면에서는 숫자로 표시합니다.",
    "comment_count": "댓글 수입니다. 서버가 자동 계산하며 목록 카드에서 표시할 수 있습니다.",
    "view_count": "조회수입니다. 인기순/랭킹 계산에 활용될 수 있습니다.",
    "favorite_count": "찜한 사람 수입니다. 디자인/샵 카드에서 인기도 지표로 표시할 수 있습니다.",
    "rating_avg": "샵 리뷰 별점의 평균값입니다. 리뷰 작성/수정/삭제 시 서버가 다시 계산합니다.",
    "rating_count": "샵에 달린 리뷰 개수입니다. 평균 별점 옆에 표시할 수 있습니다.",
    "business_hours": "요일별 영업시간입니다. 예약 가능한 시간 계산의 기준이 됩니다.",
    "refund_policy": "향후 결제/환불 기능에서 사용할 취소 시점별 환불 기준입니다. MVP에서는 사용하지 않습니다.",
    "deposit_rate": "향후 예약금 결제 기능에서 사용할 예약금 비율입니다. MVP에서는 사용하지 않습니다.",
    "auto_accept": "예약 생성 후 사장님 수락 없이 바로 확정할지 여부입니다.",
    "start_datetime": "예약 시작 시각입니다. 서버는 UTC로 저장하고 화면은 샵 로컬 시간대로 표시합니다.",
    "end_datetime": "예약 종료 시각입니다. 시작 시각과 시술 시간을 기준으로 서버가 계산합니다.",
    "duration_minutes": "시술에 걸리는 시간입니다. 예약 슬롯 계산의 핵심 기준입니다.",
    "total_price": "예약 당시 확정된 총 가격입니다. 디자인 가격이 나중에 바뀌어도 기존 예약에는 영향을 주지 않습니다.",
    "deposit_amount": "향후 예약금 결제 기능에서 사용할 예약금 금액입니다. MVP에서는 사용하지 않습니다.",
    "currency": "향후 결제 기능에서 사용할 통화입니다. 국내 서비스에서는 보통 KRW입니다.",
    "provider": "향후 결제 기능에서 사용할 PG사 이름입니다.",
    "provider_payment_id": "향후 결제 기능에서 PG사가 내려주는 결제 식별값입니다.",
    "provider_refund_id": "향후 환불 기능에서 PG사가 내려주는 환불 식별값입니다.",
    "reason": "취소, 신고 등 어떤 조치가 발생한 이유입니다.",
    "reason_detail": "신고나 처리 사유를 더 자세히 적는 선택 입력값입니다.",
    "cancel_reason": "예약 취소 또는 거절 사유입니다. 유저에게 노출할지 여부를 확인해야 합니다.",
    "failed_reason": "분석 실패나 외부 연동 실패 원인입니다. 사용자에게 보여줄 문구와 내부 로그를 분리해야 합니다.",
    "sort_order": "이미지나 항목의 노출 순서입니다. 0번을 대표 이미지로 쓰는 식의 규칙이 필요합니다.",
    "is_active": "현재 사용 가능한 대상인지 나타냅니다. 삭제 대신 비활성화할 때 사용합니다.",
}

ENTITY_FIELD_MEANINGS = {
    "User.auth_method": "로그인 제공자입니다. 현재는 Apple Sign In만 가정합니다.",
    "User.apple_user_id": "Apple이 유저별로 내려주는 고유 식별값입니다. 같은 유저의 재로그인 식별에 사용합니다.",
    "User.bio": "프로필에 표시되는 짧은 자기소개입니다.",
    "User.preferred_tags": "유저가 선호하는 네일 스타일 태그입니다. 추천/온보딩에 활용할 수 있습니다.",
    "User.device_token": "APNs 푸시 발송에 필요한 기기 토큰입니다. 유저가 직접 입력하지 않습니다.",
    "User.app_version": "앱 강제 업데이트나 호환성 판단에 쓰는 클라이언트 버전입니다.",
    "Owner.owner_name": "사업자 또는 사장님 이름입니다. 계정/사업자 인증 화면과 연결됩니다.",
    "Owner.business_number": "사업자등록번호입니다. 형식 검증과 외부 검증 여부를 정해야 합니다.",
    "Owner.business_license_image": "사업자등록증 파일입니다. 이미지 또는 PDF 업로드 정책이 필요합니다.",
    "Owner.verification_status": "사업자 인증 진행 상태입니다. 승인 대기/승인/반려 화면에 사용합니다.",
    "Shop.address": "샵의 기본 주소입니다. 주소 검색 API와 좌표 변환의 입력값입니다.",
    "Shop.address_detail": "건물명, 층, 호수 같은 상세주소입니다.",
    "Shop.lat": "샵 위치의 위도입니다. 지도/거리순 정렬에 사용하며 유저가 직접 입력하지 않습니다.",
    "Shop.lng": "샵 위치의 경도입니다. 지도/거리순 정렬에 사용하며 유저가 직접 입력하지 않습니다.",
    "Designer.career_years": "디자이너 경력 연수입니다. 프로필 표시용입니다.",
    "Designer.rank": "원장, 실장, 주니어 같은 직급입니다. 자유 입력인지 선택값인지 정해야 합니다.",
    "Designer.specialty_tags": "디자이너가 잘하는 스타일 태그입니다. 사장님 웹에서 선택하게 할 수 있습니다.",
    "Design.base_price": "디자인의 기본 가격입니다. 예약 시 total_price로 복사됩니다.",
    "Design.available_designer_ids": "이 디자인을 시술할 수 있는 디자이너 목록입니다. 자동 배정 후보가 됩니다.",
    "Design.owner_tags": "사장님이 직접 붙이는 디자인 태그입니다. AI 태그와 함께 검색에 쓰입니다.",
    "Design.ai_tags": "LLM이 이미지 분석으로 생성한 태그입니다. 사장님이 직접 입력하지 않습니다.",
    "Design.ai_color_palette": "LLM이 추출한 주요 색상 목록입니다. 검색/필터/추천에 활용할 수 있습니다.",
    "Design.ai_style_category": "LLM이 분류한 대표 스타일 카테고리입니다.",
    "DesignImage.original_url": "사장님이 올린 원본 이미지 URL입니다. LLM 재분석과 증빙을 위해 보관합니다.",
    "DesignImage.cropped_url": "LLM이 네일 영역을 추출해 만든 노출용 이미지 URL입니다.",
    "DesignImage.ai_transform_status": "LLM 1단계 네일 추출 처리 상태입니다.",
    "DesignImage.ai_classify_status": "LLM 2단계 태그 분류 처리 상태입니다.",
    "Reservation.assigned_by": "디자이너가 유저 선택으로 정해졌는지 자동 배정으로 정해졌는지 구분합니다.",
    "Reservation.reservation_policy_snapshot": "예약 당시 유저에게 보여준 취소/노쇼/변경 안내입니다. 이후 샵 문구가 바뀌어도 이 예약에는 당시 안내를 기준으로 표시합니다.",
    "Reservation.idempotency_key": "중복 예약 요청을 막기 위해 클라이언트가 보내는 고유 키입니다.",
    "Reservation.user_request_memo": "유저가 예약 시 남기는 요청사항입니다.",
    "Reservation.cancelled_at": "예약이 취소된 시각입니다.",
    "Reservation.completed_at": "시술 완료 처리된 시각입니다.",
    "Reservation.no_show_at": "노쇼 처리된 시각입니다.",
    "Payment.amount_total": "향후 결제 기능에서 PG에 실제로 결제 요청할 금액입니다.",
    "Payment.paid_at": "향후 결제 기능에서 결제가 성공한 시각입니다.",
    "Payment.failed_at": "향후 결제 기능에서 결제가 실패한 시각입니다.",
    "Payment.raw_payload_ref": "향후 결제 기능에서 PG 원문 응답 로그 위치입니다. 장애 대응용이며 화면에 노출하지 않습니다.",
    "Refund.requested_by_type": "향후 환불 기능에서 환불을 유발한 주체입니다.",
    "Refund.requested_by_id": "향후 환불 기능에서 환불을 유발한 유저/사장님/관리자 ID입니다.",
    "Refund.refund_amount": "향후 환불 기능에서 실제 환불할 금액입니다.",
    "Refund.refund_rate": "향후 환불 기능에서 적용할 환불 비율입니다.",
    "Refund.requested_at": "향후 환불 기능에서 환불 요청이 만들어진 시각입니다.",
    "Refund.completed_at": "향후 환불 기능에서 환불이 완료된 시각입니다.",
    "IdempotencyKey.key": "클라이언트가 보낸 중복 방지 키입니다. 프론트는 예약 생성마다 새 키를 만들어야 합니다.",
    "IdempotencyKey.scope": "이 키가 어떤 작업에 쓰였는지 나타냅니다. 예: 예약 생성.",
    "IdempotencyKey.actor_type": "요청을 보낸 주체 유형입니다.",
    "IdempotencyKey.actor_id": "요청을 보낸 주체의 ID입니다.",
    "IdempotencyKey.request_hash": "요청 body가 같은지 확인하기 위한 해시입니다.",
    "IdempotencyKey.response_status_code": "첫 요청의 응답 HTTP status입니다. 동일 요청 재응답에 사용합니다.",
    "IdempotencyKey.response_body_snapshot": "첫 요청의 응답 body입니다. 동일 요청 재응답에 사용합니다.",
    "IdempotencyKey.locked_until": "처리 중인 요청의 잠금 만료 시각입니다.",
    "IdempotencyKey.expires_at": "중복 방지 키를 보관하는 만료 시각입니다.",
    "Snap.author_user_id": "스네일을 작성한 유저 ID입니다. 스네일은 일반 유저가 작성합니다.",
    "Snap.caption": "스네일 본문입니다. 선택 입력이지만 최대 길이와 금칙어 정책이 필요합니다.",
    "Snap.tagged_shop_id": "스네일에 태그된 샵입니다. 샵 상세에 노출될 수 있습니다.",
    "Snap.tagged_design_id": "스네일에 태그된 디자인입니다. 디자인 상세에 노출될 수 있습니다.",
    "Snap.tagged_designer_id": "스네일에 태그된 디자이너입니다. 샵/디자인 태그와 함께 쓰입니다.",
    "Snap.tagged_reservation_id": "본인의 완료 예약을 태그한 경우 인증 뱃지 판단에 사용합니다.",
    "Snap.popularity_score": "랭킹 탭 정렬에 쓰는 서버 계산 점수입니다.",
    "Comment.parent_comment_id": "대댓글이면 부모 댓글 ID가 들어가고, 일반 댓글이면 비어 있습니다.",
    "Comment.author_type": "댓글 작성자가 일반 유저인지 샵 계정인지 구분합니다.",
    "Comment.author_user_id": "일반 유저 댓글일 때 작성자 ID입니다.",
    "Comment.author_shop_id": "샵 계정 댓글일 때 작성 샵 ID입니다.",
    "Comment.content": "댓글 본문입니다. 최대 길이와 신고/금칙어 정책이 필요합니다.",
    "Review.author_user_id": "리뷰를 작성한 유저 ID입니다. 예약자와 일치해야 합니다.",
    "Review.rating": "샵에 대한 별점입니다. 디자이너 별점은 저장하지 않습니다.",
    "Review.content": "리뷰 본문입니다. 최소/최대 길이를 정해야 합니다.",
    "Review.shop_reply": "사장님이 리뷰에 남기는 답변입니다.",
    "Review.shop_reply_at": "사장님 답변 작성 시각입니다.",
    "Report.reporter_id": "신고한 사람의 ID입니다. 유저 또는 사장님일 수 있습니다.",
    "Report.target_type": "신고 대상 종류입니다. 스네일, 댓글, 리뷰, 유저, 샵 중 하나입니다.",
    "Report.target_id": "신고 대상의 실제 ID입니다.",
    "Report.reason_code": "신고 사유 선택값입니다. 프론트/UI가 문구를 확정해야 합니다.",
    "Report.resolved_action": "운영자가 신고 처리 후 적용한 조치입니다.",
}


def deep_merge(base, update):
    for key, value in update.items():
        if isinstance(value, dict) and isinstance(base.get(key), dict):
            deep_merge(base[key], value)
        else:
            base[key] = value
    return base


def extract_spec_data_blocks(path):
    text = path.read_text(encoding="utf-8")
    blocks = []
    marker = "```json spec-data"
    start = 0
    while True:
        marker_pos = text.find(marker, start)
        if marker_pos == -1:
            break
        block_start = text.find("\n", marker_pos)
        if block_start == -1:
            break
        block_end = text.find("```", block_start + 1)
        if block_end == -1:
            raise ValueError(f"닫히지 않은 spec-data 코드블록: {path}")
        raw = text[block_start + 1:block_end].strip()
        if raw:
            try:
                blocks.append(json.loads(raw))
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSON 파싱 실패: {path} line {exc.lineno}, col {exc.colno}: {exc.msg}") from exc
        start = block_end + 3
    return blocks


def load_spec_data():
    with CANONICAL_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if SPEC_TEXT_DIR.exists():
        for path in sorted(SPEC_TEXT_DIR.glob("*.md")):
            for block in extract_spec_data_blocks(path):
                deep_merge(data, block)
    return data


def set_widths(ws, values):
    for i, width in enumerate(values, start=1):
        letter = get_column_letter(i)
        current = ws.column_dimensions[letter].width or DEFAULT_COL_WIDTH
        ws.column_dimensions[letter].width = min(max(current, width), MAX_COL_WIDTH)


def display_units(value):
    text = "" if value is None else str(value)
    units = 0.0
    for char in text:
        if char == "\n":
            continue
        units += 1.65 if ord(char) > 127 else 1.0
    return units


def estimate_line_count(value, width):
    text = "" if value is None else str(value)
    if not text:
        return 1
    capacity = max(8, int(width * 1.05))
    lines = 0
    for part in text.splitlines() or [""]:
        units = display_units(part)
        lines += max(1, int((units + capacity - 1) // capacity))
    return lines


def row_height_for_values(values, widths):
    line_count = 1
    for idx, value in enumerate(values):
        width = widths[idx] if idx < len(widths) else DEFAULT_COL_WIDTH
        line_count = max(line_count, estimate_line_count(value, width))
    return min(MAX_ROW_HEIGHT, max(MIN_ROW_HEIGHT, line_count * LINE_HEIGHT + ROW_PADDING))


def merged_width(ws, start_col, end_col):
    total = 0
    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        total += ws.column_dimensions[letter].width or DEFAULT_COL_WIDTH
    return max(total, DEFAULT_COL_WIDTH)


def effective_cell_width(ws, cell):
    for cell_range in ws.merged_cells.ranges:
        if (
            cell.row == cell_range.min_row
            and cell.column == cell_range.min_col
        ):
            return merged_width(ws, cell_range.min_col, cell_range.max_col)
        if cell.coordinate in cell_range:
            return 0
    return ws.column_dimensions[get_column_letter(cell.column)].width or DEFAULT_COL_WIDTH


def row_height_for_cells(ws, cells):
    line_count = 1
    for cell in cells:
        if cell.value is None:
            continue
        width = effective_cell_width(ws, cell)
        if width <= 0:
            continue
        line_count = max(line_count, estimate_line_count(cell.value, width))
    return min(MAX_ROW_HEIGHT, max(MIN_ROW_HEIGHT, line_count * LINE_HEIGHT + ROW_PADDING))


def title(ws, text, subtitle=None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = TL
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUB_FONT
        ws["A2"].alignment = TL
        ws.row_dimensions[2].height = 22


def section(ws, row, text, cols=7):
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=1).alignment = TL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = BORDER
    ws.row_dimensions[row].height = 24


def note(ws, row, text, cols=7, fill=NOTE_FILL, height=48):
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = BODY
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = TL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = BORDER
    auto_height = row_height_for_values([text], [merged_width(ws, 1, cols)])
    ws.row_dimensions[row].height = max(height, auto_height)


def table(ws, start_row, headers, rows, review_cols=None, widths=None):
    review_cols = review_cols or []
    if widths:
        set_widths(ws, widths)
    effective_widths = [
        ws.column_dimensions[get_column_letter(i)].width or DEFAULT_COL_WIDTH
        for i in range(1, len(headers) + 1)
    ]
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[start_row].height = row_height_for_values(headers, effective_widths)
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = BODY
            cell.alignment = TL
            cell.border = BORDER
            if c_idx in review_cols:
                cell.fill = REQUIRED_FILL
        ws.row_dimensions[r_idx].height = row_height_for_values(row, effective_widths)
    return start_row + len(rows) + 1


def finalize_workbook(wb):
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "A4"
        ws.sheet_format.defaultRowHeight = MIN_ROW_HEIGHT
        for row in ws.iter_rows():
            estimated = row_height_for_cells(ws, row)
            current = ws.row_dimensions[row[0].row].height or MIN_ROW_HEIGHT
            ws.row_dimensions[row[0].row].height = max(current, estimated)
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="top",
                        wrap_text=True,
                    )


def page_guide_text(data, key):
    guide = data.get("page_guides", {}).get(key)
    if not guide:
        return None
    return (
        f"이 페이지에서 다루는 것: {guide['covers']}\n"
        f"관련 업무: {guide['related_work']}\n"
        f"읽는 방법: {guide['how_to_use']}"
    )


def add_page_guide(ws, data, key, row):
    guide_text = page_guide_text(data, key)
    if not guide_text:
        return row
    note(ws, row, guide_text, fill=WHO_FILL, height=76)
    return row + 2


def required_help(required):
    if required == "자동":
        return "서버가 자동으로 만들거나 계산합니다. 입력 UI를 만들 필요는 없습니다."
    if required == "필수":
        return "사용자 또는 사장님이 반드시 제공해야 합니다. 누락 시 화면 에러 문구가 필요합니다."
    if required == "선택":
        return "입력하지 않을 수 있습니다. 비어 있을 때의 기본 표시/빈 상태를 정해야 합니다."
    if required == "조건부":
        return "특정 조건에서만 필요합니다. 어떤 조건에서 보여주고 검증할지 화면 흐름을 확인해야 합니다."
    return "필수 여부가 특수한 값입니다. 화면 흐름에 맞는 처리 방식을 확인해주세요."


def field_meaning(entity, name, note_text):
    return (
        ENTITY_FIELD_MEANINGS.get(f"{entity}.{name}")
        or COMMON_FIELD_MEANINGS.get(name)
        or note_text
        or "서버가 업무 흐름을 처리하기 위해 사용하는 값입니다. 화면 노출/입력 여부를 협업자가 확인해야 합니다."
    )


def collaborator_prompt(entity, name, required):
    if name.endswith("_id") or name in {"key", "request_hash", "raw_payload_ref"}:
        return "대부분 화면에 직접 보여주지 않는 내부 식별값입니다. 상세 화면 이동, 딥링크, 디버깅에 필요한지만 확인해주세요."
    if required == "자동":
        return "직접 입력받지 않습니다. 목록/상세/상태 표시 또는 정렬·필터에 필요한지 확인해주세요."
    if required == "필수":
        return "입력 UI, placeholder, 누락 시 에러 문구, 최대/최소 길이 또는 선택 옵션을 정해주세요."
    if required == "선택":
        return "입력하지 않았을 때 숨길지, 빈 문구를 보여줄지, 나중에 수정 가능하게 할지 정해주세요."
    if required == "조건부":
        return "어떤 상태/선택/권한에서 필수로 바뀌는지 화면 조건을 정해주세요."
    return "화면에서 입력·표시·수정 가능 여부를 확인해주세요."


def field_rows(fields, entity, annotations=None):
    annotations = annotations or {}
    entity_notes = annotations.get(entity, {}) if annotations else {}
    return [
        [
            name,
            "",
            typ,
            required,
            field_meaning(entity, name, note_text),
            required_help(required),
            collaborator_prompt(entity, name, required),
            entity_notes.get(name, ""),
        ]
        for name, typ, required, note_text in fields
    ]


def api_call_context(purpose):
    return f"'{purpose}' 화면/동작에서 호출됩니다. 실제 화면명, 버튼명, 탭명, 상태명을 더 구체적으로 적어주세요."


def api_frontend_check(params):
    return f"요청값({params})을 프론트/웹에서 준비할 수 있는지, 성공 후 화면 이동과 실패 문구가 필요한지 확인해주세요."


def add_field_sheet(wb, name, sheet_title, subtitle, entities, data):
    ws = wb.create_sheet(name)
    title(ws, sheet_title, subtitle)
    r = 4
    r = add_page_guide(ws, data, name, r)
    for entity in entities:
        section(ws, r, f"{entity} 필드 정의")
        r += 1
        r = table(
            ws,
            r,
            ["백엔드 필드명", "화면 라벨/표현", "타입", "필수?", "필드 의미", "입력/생성 방식", "협업자 확인 사항", "추가 요청/결정"],
            field_rows(data["entities"][entity], entity, data.get("collaborator_annotations", {}).get("entities", {})),
            review_cols=[2, 7, 8],
            widths=[22, 20, 14, 10, 42, 36, 42, 28],
        )
        r += 2


def add_api_sheet(wb, name, sheet_title, subtitle, api_groups, data, note_text=None):
    ws = wb.create_sheet(name)
    title(ws, sheet_title, subtitle)
    r = 4
    r = add_page_guide(ws, data, name, r)
    if note_text:
        note(ws, r, note_text)
        r += 2
    for label, key in api_groups:
        section(ws, r, label)
        r += 1
        rows = [
            [endpoint, purpose, params, api_call_context(purpose), api_frontend_check(params), ""]
            for endpoint, purpose, params in data["apis"][key]
        ]
        r = table(
            ws,
            r,
            ["엔드포인트", "용도", "요청값", "호출 화면/상황", "프론트 확인 사항", "추가 요청/결정"],
            rows,
            review_cols=[4, 5, 6],
            widths=[32, 24, 34, 36, 44, 28],
        )
        r += 2


def append_api_groups(ws, api_groups, data):
    r = ws.max_row + 2
    for label, key in api_groups:
        section(ws, r, label)
        r += 1
        rows = [
            [endpoint, purpose, params, api_call_context(purpose), api_frontend_check(params), ""]
            for endpoint, purpose, params in data["apis"][key]
        ]
        r = table(
            ws,
            r,
            ["엔드포인트", "용도", "요청값", "호출 화면/상황", "프론트 확인 사항", "추가 요청/결정"],
            rows,
            review_cols=[4, 5, 6],
            widths=[32, 24, 34, 36, 44, 28],
        )
        r += 2


def append_field_groups(ws, entities, data):
    r = ws.max_row + 2
    for entity in entities:
        section(ws, r, f"{entity} 필드 정의")
        r += 1
        r = table(
            ws,
            r,
            ["백엔드 필드명", "화면 라벨/표현", "타입", "필수?", "필드 의미", "입력/생성 방식", "협업자 확인 사항", "추가 요청/결정"],
            field_rows(data["entities"][entity], entity, data.get("collaborator_annotations", {}).get("entities", {})),
            review_cols=[2, 7, 8],
            widths=[22, 20, 14, 10, 42, 36, 42, 28],
        )
        r += 2


def append_reservation_statuses(ws, data):
    r = ws.max_row + 2
    section(ws, r, "예약 상태(status) 정의")
    r += 1
    table(ws, r, ["status", "의미", "트리거", "유저 표시"], data["reservation_statuses"])


def append_reservation_transitions(ws, data):
    r = ws.max_row + 2
    section(ws, r, "예약 상태 전이표")
    r += 1
    rows = [
        [
            item["current_status"],
            item["actor"],
            item["action"],
            item["api"],
            item["next_status"],
            item.get("process_note", item.get("payment_refund", "")),
            item["collaborator_check"],
        ]
        for item in data["reservation_state_transitions"]
    ]
    table(
        ws,
        r,
        ["현재 상태", "주체", "액션", "API", "다음 상태", "처리 내용", "협업자 확인"],
        rows,
        review_cols=[7],
        widths=[18, 14, 24, 34, 20, 28, 30],
    )


def append_availability_rules(ws, data):
    rules = data["availability_rules"]
    r = ws.max_row + 2
    section(ws, r, "가용 슬롯 계산 규칙")
    r += 1
    note(ws, r, rules["purpose"], height=34)
    r += 2
    rows = [["백엔드 기본 규칙", item, ""] for item in rules["backend_default_rules"]]
    rows += [["팀 내부 결정 필요", item, ""] for item in rules["team_decisions_needed"]]
    rows += [["협업자 확인", item, ""] for item in rules["collaborator_checks"]]
    table(ws, r, ["구분", "내용", "확정/의견"], rows, review_cols=[3], widths=[24, 80, 34])


def add_readme(wb, data):
    ws = wb.active
    ws.title = "1.README"
    title(ws, data["document"]["title"], data["document"]["subtitle"])
    r = 4
    r = add_page_guide(ws, data, "1.README", r)
    section(ws, r, "현재 명세의 우선 목표", 2)
    r += 1
    ws.cell(row=r, column=1, value="목표").font = BODY_B
    ws.cell(row=r, column=2, value=data["collaboration_goal"]["primary_goal"]).font = BODY
    ws.cell(row=r, column=2).alignment = TL
    r += 1
    ws.cell(row=r, column=1, value="아님").font = BODY_B
    ws.cell(row=r, column=2, value=data["collaboration_goal"]["not_primary_goal"]).font = BODY
    ws.cell(row=r, column=2).alignment = TL
    r += 2
    section(ws, r, "협업자별 작성 책임", 2)
    r += 1
    for role, responsibility in data["collaboration_goal"]["role_responsibilities"]:
        ws.cell(row=r, column=1, value=role).font = BODY_B
        ws.cell(row=r, column=2, value=responsibility).font = BODY
        ws.cell(row=r, column=2).alignment = TL
        r += 1
    r += 1
    section(ws, r, "v3 변경사항", 2)
    r += 1
    for item in data["changes"]:
        ws.cell(row=r, column=1, value="•").font = BODY_B
        ws.cell(row=r, column=2, value=item).font = BODY
        ws.cell(row=r, column=2).alignment = TL
        r += 1
    r += 1
    section(ws, r, "사용 원칙", 2)
    r += 1
    for item in [
        "노란색 칸은 데이터 값이 아니라 데이터 정의/스펙을 적는 칸입니다.",
        "탐색과 수정 논의는 spec_canonical 폴더의 JSON/Markdown을 우선합니다.",
        "엑셀은 최종 공유 산출물로 재생성합니다.",
    ]:
        ws.cell(row=r, column=1, value="•").font = BODY_B
        ws.cell(row=r, column=2, value=item).font = BODY
        ws.cell(row=r, column=2).alignment = TL
        r += 1
    r += 1
    section(ws, r, "시스템 주체", 2)
    r += 1
    for actor in data["actors"]:
        ws.cell(row=r, column=1, value=actor["name"]).font = BODY_B
        ws.cell(row=r, column=2, value=actor["description"]).font = BODY
        ws.cell(row=r, column=2).alignment = TL
        r += 1
    set_widths(ws, [28, 100])


def add_flow(wb, data):
    ws = wb.create_sheet("2.전체플로우")
    title(ws, "전체 데이터 플로우", "주체별로 누가 언제 무엇을 백엔드에 보내는지")
    r = 4
    r = add_page_guide(ws, data, "2.전체플로우", r)
    note(ws, r, "주체: 유저 앱, 사장님 웹, LLM 서비스, 카카오 알림톡, 운영자/어드민", height=30)
    rows = [
        ["1", "사장님 웹", "회원가입 + 샵 등록", "샵/디자이너/스케줄 INSERT"],
        ["2", "사장님 웹", "디자인 사진 업로드 + 디자이너 선택", "S3 저장 후 LLM 2단계 분석"],
        ["3a", "LLM Transform", "네일 영역 추출", "원본 → 규격화 이미지"],
        ["3b", "LLM Classification", "태그/색상/스타일 분류", "검색용 AI 메타데이터 생성"],
        ["4", "유저 앱", "Apple Sign In", "유저 생성 + JWT 발급"],
        ["5", "유저 앱", "DB 기반 자연어형 검색/디자인 상세", "디자인 + 스네일 + 리뷰 반환"],
        ["6", "유저 앱", "결제 없는 예약 요청 생성", "중복 요청 방지 후 사장님 카카오 알림톡 발송"],
        ["7", "사장님 웹", "예약 수락/거절", "상태 업데이트 + 유저 APNs"],
        ["8", "유저 앱", "완료 후 리뷰/스네일", "리뷰는 샵 별점 반영, 스네일은 커뮤니티 지표로만 사용"],
    ]
    table(ws, r + 2, ["#", "주체", "행동", "백엔드 처리"], rows, widths=[6, 22, 36, 52])


def add_statuses(wb, data):
    ws = wb.create_sheet("4-예약상태")
    title(ws, "예약 상태 정의", "예약 status와 트리거")
    table(ws, 4, ["status", "의미", "트리거", "유저 표시"], data["reservation_statuses"], widths=[22, 24, 40, 32])


def add_llm(wb, data):
    ws = wb.create_sheet("12.LLM명세")
    title(ws, "LLM 네일 분석 API 명세", "Visual Transformer + Classification 2단계")
    r = 4
    r = add_page_guide(ws, data, "12.LLM명세", r)
    note(ws, r, "LLM 명세는 가장 중요한 협업 지점입니다. 태그 사전은 검색 품질과 직접 연결됩니다.", fill=WARN_FILL, height=48)
    r += 2
    llm_input = data.get("llm_collaborator_input", {})
    for key, label in [("transform", "1단계 Transform"), ("classify", "2단계 Classification")]:
        spec = data["llm"][key]
        section_input = llm_input.get(key, {})
        section(ws, r, label)
        r += 1
        rows = [
            ["목적", spec["purpose"], section_input.get("purpose", "")],
            ["제안 endpoint", spec["suggested_endpoint"], section_input.get("suggested_endpoint", "")],
            ["Request fields", ", ".join(spec["request_fields"]), section_input.get("request_fields", "")],
            ["Response fields", ", ".join(spec["response_fields"]), section_input.get("response_fields", "")],
        ]
        if "recommendation" in spec:
            rows.append(["운영 권장", spec["recommendation"], section_input.get("recommendation", "")])
        r = table(ws, r, ["항목", "내용", "LLM 측 확정값"], rows, review_cols=[3], widths=[22, 84, 36])
        r += 2
    section(ws, r, "표준 태그 사전")
    r += 1
    tag_input = llm_input.get("standard_tags", {})
    tag_rows = [[k, ", ".join(v), tag_input.get(k, "")] for k, v in data["llm"]["standard_tags"].items()]
    r = table(ws, r, ["카테고리", "백엔드 제안", "LLM 측 확정"], tag_rows, review_cols=[3], widths=[22, 92, 36])
    r += 2
    section(ws, r, "에러 코드")
    r += 1
    err_input = llm_input.get("error_codes", {})
    err_rows = [row + [err_input.get(row[0], "")] for row in data["llm"]["error_codes"]]
    r = table(ws, r, ["코드", "상황", "사장님 노출 메시지", "LLM 측 수정"], err_rows, review_cols=[4], widths=[20, 36, 70, 24])
    r += 2
    section(ws, r, "LLM 작업자 작성 필요 질문")
    r += 1
    rows = [[idx, question, ""] for idx, question in enumerate(data["llm"]["worker_questions"], start=1)]
    table(ws, r, ["#", "질문", "LLM 측 답변"], rows, review_cols=[3], widths=[6, 90, 40])


def add_notifications(wb, data):
    ws = wb.create_sheet("13.알림")
    title(ws, "알림 명세", "유저 앱 APNs + 사장님 카카오 알림톡")
    r = 4
    r = add_page_guide(ws, data, "13.알림", r)
    note(ws, r, "사장님은 카카오 알림톡을 사용하고, 웹푸시는 MVP에서 제외합니다.", fill=WARN_FILL, height=36)
    table(ws, r + 2, ["알림 종류", "수신자", "채널"], data["notifications"], widths=[36, 20, 24])


def add_decisions(wb, data):
    ws = wb.create_sheet("14.의사결정기록")
    title(ws, "의사결정 기록", "확정된 정책 모음")
    r = 4
    r = add_page_guide(ws, data, "14.의사결정기록", r)
    rows = [[idx, topic, decision] for idx, (topic, decision) in enumerate(data["decisions"], start=1)]
    r = table(ws, r, ["#", "주제", "결정 내용"], rows, widths=[6, 28, 90])
    r += 2
    section(ws, r, "팀 내부 결정 필요 사항", 4)
    r += 1
    pending_rows = [
        [item["topic"], item["decision_needed"], item["owner"], item.get("answer", "")]
        for item in data["internal_decisions_needed"]
    ]
    table(ws, r, ["주제", "결정 필요", "담당", "결정/메모"], pending_rows, review_cols=[4], widths=[26, 72, 18, 30])


def add_common_rules(wb, data):
    ws = wb.create_sheet("16.공통_API권한")
    title(ws, "공통 API / 인증 권한 규칙", "프론트 협업을 매끄럽게 하기 위한 공통 약속")
    r = 4
    r = add_page_guide(ws, data, "16.공통_API권한", r)
    section(ws, r, "권한 주체")
    r += 1
    r = table(ws, r, ["actor", "이름", "설명"], data["auth_permission_model"]["actors"], widths=[18, 24, 80])
    r += 2
    section(ws, r, "권한 표")
    r += 1
    r = table(
        ws,
        r,
        ["기능", "비로그인", "로그인 유저", "사장님 Owner", "어드민", "협업자 확인"],
        data["auth_permission_model"]["permission_matrix"],
        review_cols=[6],
        widths=[28, 18, 24, 26, 18, 34],
    )
    r += 2
    section(ws, r, "공통 에러 코드")
    r += 1
    r = table(
        ws,
        r,
        ["코드", "의미", "프론트 처리/문구"],
        [row + [""] for row in data["common_api_rules"]["error_response"]["codes"]],
        review_cols=[3],
        widths=[26, 54, 44],
    )
    r += 2
    section(ws, r, "페이지네이션 / Rate Limit / 빈 상태")
    r += 1
    rows = []
    for item in data["common_api_rules"]["pagination"]:
        rows.append(["페이지네이션", item, ""])
    for item in data["common_api_rules"]["rate_limit"]:
        rows.append(["Rate Limit", item, ""])
    for item in data["common_api_rules"]["empty_state"]:
        rows.append(["빈 상태", item, ""])
    for item in data["common_api_rules"]["time_format"]:
        rows.append(["시간 포맷", item, ""])
    table(ws, r, ["구분", "규칙", "프론트 확인"], rows, review_cols=[3], widths=[22, 80, 34])


def add_checklist(wb, data):
    ws = wb.create_sheet("15.체크리스트")
    title(ws, "최종 체크리스트", "회신 전 점검")
    r = 4
    r = add_page_guide(ws, data, "15.체크리스트", r)
    rows = []
    for owner, items in data["checklist"].items():
        rows.extend([[owner, item, "", ""] for item in items])
    table(ws, r, ["담당", "항목", "상태", "메모/막힌 점"], rows, review_cols=[3, 4], widths=[18, 70, 16, 34])
    dv = DataValidation(type="list", formula1='"예,아니오,미정"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C{r + 1}:C{r + len(rows)}")


def main():
    data = load_spec_data()

    wb = Workbook()
    add_readme(wb, data)
    add_flow(wb, data)
    add_field_sheet(wb, "3.유저(앱)_회원", "유저(앱) - 회원가입 / 로그인 / 프로필", "iOS 유저 앱 클라이언트", ["User"], data)
    add_api_sheet(
        wb,
        "4.유저(앱)_탐색예약",
        "유저(앱) - 검색 / 피드 / 예약",
        "iOS 유저 앱 클라이언트",
        [("검색 / 피드 API", "search"), ("예약 API", "reservation")],
        data,
        "검색 결과 0건 시 유사 디자인 추천을 포함합니다. 예약 생성에는 Idempotency-Key가 필요합니다.",
    )
    append_field_groups(wb["4.유저(앱)_탐색예약"], ["Reservation", "IdempotencyKey"], data)
    append_reservation_statuses(wb["4.유저(앱)_탐색예약"], data)
    append_reservation_transitions(wb["4.유저(앱)_탐색예약"], data)
    append_availability_rules(wb["4.유저(앱)_탐색예약"], data)
    add_field_sheet(wb, "5.사장님(웹)_샵관리", "사장님(웹) - 회원가입 / 샵 / 디자이너 관리", "네일샵 사장님 전용 웹 클라이언트", ["Owner", "Shop", "Designer"], data)
    append_api_groups(
        wb["5.사장님(웹)_샵관리"],
        [("Owner/Auth API", "owner_auth"), ("Shop API", "owner_shop"), ("Designer API", "owner_designer")],
        data,
    )
    add_field_sheet(wb, "6.사장님(웹)_디자인", "사장님(웹) - 디자인 등록 / 관리", "사진 업로드 → LLM 2단계 분석", ["Design", "DesignImage"], data)
    append_api_groups(wb["6.사장님(웹)_디자인"], [("Design API", "owner_design")], data)
    add_api_sheet(wb, "7.사장님(웹)_예약", "사장님(웹) - 예약 관리", "예약 목록 조회, 수락/거절, 캘린더 뷰", [("예약 관리 API", "owner_reservation")], data)
    add_field_sheet(wb, "8.커뮤니티_스네일", "커뮤니티 - 스네일(Snap)", "통합 커뮤니티, 스네일/랭킹/팔로잉", ["Snap"], data)
    append_api_groups(wb["8.커뮤니티_스네일"], [("스네일 API", "snap")], data)
    add_field_sheet(wb, "9.커뮤니티_댓글", "커뮤니티 - 댓글 / 좋아요 / 팔로우", "depth 2까지 허용", ["Comment"], data)
    append_api_groups(wb["9.커뮤니티_댓글"], [("댓글/좋아요/팔로우 API", "comment_like_follow")], data)
    add_field_sheet(wb, "10.커뮤니티_리뷰", "커뮤니티 - 리뷰", "1예약=1리뷰, 샵 별점만", ["Review"], data)
    append_api_groups(wb["10.커뮤니티_리뷰"], [("리뷰 API", "review")], data)
    add_field_sheet(wb, "11.커뮤니티_신고", "커뮤니티 - 신고 / 모더레이션", "스네일, 댓글, 리뷰, 유저, 샵 신고", ["Report"], data)
    add_llm(wb, data)
    add_notifications(wb, data)
    add_decisions(wb, data)
    add_checklist(wb, data)
    add_common_rules(wb, data)
    finalize_workbook(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUTPUT_PATH)
        saved_path = OUTPUT_PATH
    except PermissionError:
        saved_path = OUTPUT_PATH.with_name(f"{OUTPUT_PATH.stem}_updated{OUTPUT_PATH.suffix}")
        try:
            wb.save(saved_path)
        except PermissionError:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            saved_path = OUTPUT_PATH.with_name(f"{OUTPUT_PATH.stem}_{stamp}{OUTPUT_PATH.suffix}")
            wb.save(saved_path)
    print(f"saved: {saved_path}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
