"""
Cell-by-cell diff between the original workbook (canonical export) and the
team-filled workbook. Outputs a structured diff report we can review before
deciding what to merge back into canonical JSON.
"""
import sys
import io
import json
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
OLD = ROOT / "outputs" / "네일예약_백엔드_협업명세서_v3.xlsx"
NEW = ROOT / "outputs" / "snail_백엔드_협업명세서_v3_20260527_1825.xlsx"

def cells(ws):
    out = {}
    for row in ws.iter_rows(values_only=False):
        for c in row:
            if c.value is None:
                continue
            out[c.coordinate] = c.value
    return out

wb_old = load_workbook(OLD, data_only=True)
wb_new = load_workbook(NEW, data_only=True)

report = []
summary = {"sheets": {}, "totals": {"added": 0, "changed": 0, "removed": 0}}

old_names = set(wb_old.sheetnames)
new_names = set(wb_new.sheetnames)
removed_sheets = sorted(old_names - new_names)
added_sheets = sorted(new_names - old_names)

if removed_sheets:
    report.append(f"=== Sheets removed in NEW ===")
    for s in removed_sheets:
        report.append(f"  - {s}")
if added_sheets:
    report.append(f"=== Sheets added in NEW ===")
    for s in added_sheets:
        report.append(f"  + {s}")

for name in wb_new.sheetnames:
    if name not in old_names:
        continue
    ws_old = wb_old[name]
    ws_new = wb_new[name]
    old_cells = cells(ws_old)
    new_cells = cells(ws_new)
    added = []
    changed = []
    removed = []
    keys = set(old_cells) | set(new_cells)
    for k in keys:
        ov = old_cells.get(k)
        nv = new_cells.get(k)
        if ov == nv:
            continue
        if ov is None:
            added.append((k, nv))
        elif nv is None:
            removed.append((k, ov))
        else:
            changed.append((k, ov, nv))
    summary["sheets"][name] = {
        "added": len(added), "changed": len(changed), "removed": len(removed),
    }
    summary["totals"]["added"] += len(added)
    summary["totals"]["changed"] += len(changed)
    summary["totals"]["removed"] += len(removed)

    if added or changed or removed:
        report.append(f"\n### Sheet: {name}")
        report.append(f"  added={len(added)} changed={len(changed)} removed={len(removed)}")

# Print summary first
print("=" * 70)
print("SUMMARY")
print("=" * 70)
print(f"Total added cells:   {summary['totals']['added']}")
print(f"Total changed cells: {summary['totals']['changed']}")
print(f"Total removed cells: {summary['totals']['removed']}")
print()
print(f"{'Sheet':<40} {'added':>8} {'changed':>8} {'removed':>8}")
print("-" * 70)
for name, s in summary["sheets"].items():
    if s["added"] or s["changed"] or s["removed"]:
        print(f"{name:<40} {s['added']:>8} {s['changed']:>8} {s['removed']:>8}")

# Save detailed diff to JSON for downstream tools
detail = {"sheets": {}, "removed_sheets": removed_sheets, "added_sheets": added_sheets}
for name in wb_new.sheetnames:
    if name not in old_names:
        continue
    ws_old = wb_old[name]
    ws_new = wb_new[name]
    old_cells = cells(ws_old)
    new_cells = cells(ws_new)
    keys = sorted(set(old_cells) | set(new_cells))
    sheet_diff = {"added": {}, "changed": {}, "removed": {}}
    for k in keys:
        ov = old_cells.get(k)
        nv = new_cells.get(k)
        if ov == nv:
            continue
        if ov is None:
            sheet_diff["added"][k] = str(nv)
        elif nv is None:
            sheet_diff["removed"][k] = str(ov)
        else:
            sheet_diff["changed"][k] = {"old": str(ov), "new": str(nv)}
    if any(sheet_diff.values()):
        detail["sheets"][name] = sheet_diff

out_path = ROOT / "outputs" / "_sheet_diff.json"
out_path.write_text(json.dumps(detail, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\nDetailed diff written: {out_path}")
