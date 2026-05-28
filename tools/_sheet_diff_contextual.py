"""
Build a human-readable report that, for each filled-in cell in the team's
xlsx, shows the row context (field name, label, etc.) so we can decide where
each annotation belongs in canonical JSON.
"""
import io
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
OLD = ROOT / "outputs" / "네일예약_백엔드_협업명세서_v3.xlsx"
NEW = ROOT / "outputs" / "snail_백엔드_협업명세서_v3_20260527_1825.xlsx"
DIFF = ROOT / "outputs" / "_sheet_diff.json"

wb_new = load_workbook(NEW, data_only=True)
diff = json.loads(DIFF.read_text(encoding="utf-8"))

# For field sheets we want columns A,B,C,D,E,F,G as context for whatever cell was filled
# Header row layout (from build_workbook.add_field_sheet):
# A=백엔드 필드명, B=화면 라벨, C=타입, D=필수?, E=필드 의미, F=입력/생성 방식,
# G=협업자 확인 사항, H=추가 요청/결정

FIELD_SHEETS = {
    "3.유저(앱)_회원",
    "5.사장님(웹)_샵관리",
    "6.사장님(웹)_디자인",
    "8.커뮤니티_스네일",
    "9.커뮤니티_댓글",
    "10.커뮤니티_리뷰",
    "11.커뮤니티_신고",
}

def col_letter_to_idx(coord):
    # Strip digits to get letters
    letters = "".join(ch for ch in coord if ch.isalpha())
    row = int("".join(ch for ch in coord if ch.isdigit()))
    return letters, row

def row_context(ws, row, cols="ABCDEFG"):
    out = {}
    for letter in cols:
        v = ws[f"{letter}{row}"].value
        if v is not None:
            out[letter] = str(v)
    return out

report_lines = []
report_lines.append(f"{'='*80}")
report_lines.append("CONTEXTUAL DIFF REPORT")
report_lines.append(f"{'='*80}\n")

# Group by sheet
grouped = {}

for sheet_name, sheet_diff in diff["sheets"].items():
    ws = wb_new[sheet_name]
    added = sheet_diff.get("added", {})
    if not added:
        continue
    grouped[sheet_name] = []
    # Sort by row number
    sorted_coords = sorted(added.keys(), key=lambda c: (int("".join(ch for ch in c if ch.isdigit())), c))
    for coord in sorted_coords:
        letters, row = col_letter_to_idx(coord)
        ctx = row_context(ws, row)
        grouped[sheet_name].append({
            "coord": coord,
            "row": row,
            "col": letters,
            "context": ctx,
            "value": added[coord],
        })

for sheet_name, entries in grouped.items():
    report_lines.append(f"\n{'─'*80}")
    report_lines.append(f"📋 {sheet_name}  ({len(entries)} filled cells)")
    report_lines.append(f"{'─'*80}")
    for e in entries:
        ctx = e["context"]
        col_a = ctx.get("A", "")
        col_b = ctx.get("B", "")
        col_e = ctx.get("E", "")  # 필드 의미 (for field sheets)
        col_g = ctx.get("G", "")  # 협업자 확인 사항
        report_lines.append(f"\n  [{e['coord']}] (row {e['row']}, col {e['col']})")
        if sheet_name in FIELD_SHEETS:
            if col_a:
                report_lines.append(f"    field:  {col_a}")
            if col_b:
                report_lines.append(f"    label:  {col_b}")
            if col_e:
                report_lines.append(f"    meaning:{col_e[:80]}...")
            if col_g:
                report_lines.append(f"    ask:    {col_g[:80]}...")
        else:
            for letter in "ABCDEFG":
                if letter in ctx:
                    val = ctx[letter]
                    if len(val) > 80:
                        val = val[:80] + "..."
                    report_lines.append(f"    {letter}: {val}")
        # The filled value
        val = e["value"]
        if len(val) > 200:
            val_disp = val[:200] + "..."
        else:
            val_disp = val
        report_lines.append(f"    ✏️  FILLED: {val_disp}")

# Write to file too
out = "\n".join(report_lines)
print(out)
(ROOT / "outputs" / "_sheet_diff_contextual.md").write_text(out, encoding="utf-8")
print(f"\n\nSaved: {ROOT / 'outputs' / '_sheet_diff_contextual.md'}")
